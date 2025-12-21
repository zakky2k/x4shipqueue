from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from x4shipqueue.export.schema import MATERIAL_SCHEMAS
from x4shipqueue.models import EquipmentRow, HullRow

"""
Schemas are display-only
Column names ≠ ware IDs
"""
    

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def components_to_map(
    components: List[Tuple[str, int]],
) -> Dict[str, int]:
    """
    Convert [(material, amount), ...] into a summed lookup.
    Defensive against duplicate materials.
    """
    result: Dict[str, int] = {}
    for name, qty in components:
        result[name] = result.get(name, 0) + qty
    return result


def autofit_columns(ws: Worksheet) -> None:
    """
    Approximate Excel-style auto-fit.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2


# -----------------------------------------------------------------------------
# Equipment sheets
# -----------------------------------------------------------------------------

def write_equipment_sheet(
    ws: Worksheet,
    rows: List[EquipmentRow],
    equipment_type: str,
) -> None:
    try:
        material_columns = MATERIAL_SCHEMAS[equipment_type]
    except KeyError:
        raise KeyError(
            f"No material schema defined for equipment type '{equipment_type}'"
        )

    headers = [
        "Source",
        "Equipment ID",
        "Equipment Name",
        "Race",
        "Size",
        "Mk",
        *material_columns,
    ]
    ws.append(headers)

    for r in rows:
        comp_map = components_to_map(r.production.components)

        excel_row = [
            r.source,
            r.equipment_id,
            r.equipment_name,
            r.race,
            r.size,
            r.mk,
        ]

        for mat in material_columns:
            excel_row.append(comp_map.get(mat, ""))

        ws.append(excel_row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


# -----------------------------------------------------------------------------
# Hull sheet
# -----------------------------------------------------------------------------

def write_hulls_sheet(
    ws: Worksheet,
    rows: List[HullRow],
) -> None:
    # Hulls use a fixed schema key
    material_columns = MATERIAL_SCHEMAS["Hulls"]

    headers = [
        "Source",
        "macro_id",
        "Hull ID",
        "Hull Name",
        "Faction",
        "Size",
        *material_columns,
    ]
    ws.append(headers)

    for r in rows:
        comp_map = components_to_map(r.production.components)

        excel_row = [
            r.source,
            r.macro_id,
            r.hull_id,
            r.hull_name,
            r.faction,
            r.size,
        ]

        for mat in material_columns:
            excel_row.append(comp_map.get(mat, ""))

        ws.append(excel_row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


# -----------------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------------

def export_to_excel(
    out_path: Path,
    equipment_by_category: Dict[str, List[EquipmentRow]],
    hull_rows: List[HullRow],
    include_all_equipment: bool,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    order = ["Engines", "Thrusters", "Shields", "Weapons", "Turrets"]

    for cat in order:
        ws = wb.create_sheet(cat)
        write_equipment_sheet(
            ws,
            equipment_by_category.get(cat, []),
            equipment_type=cat,
        )

    if include_all_equipment:
        all_rows: List[EquipmentRow] = []
        for cat in order:
            all_rows.extend(equipment_by_category.get(cat, []))

        ws = wb.create_sheet("All_Equipment")
        # NOTE: mixed-category sheet; schema choice is intentional
        write_equipment_sheet(
            ws,
            all_rows,
            equipment_type="Weapons",
        )

    ws = wb.create_sheet("Hulls")
    write_hulls_sheet(ws, hull_rows)

    wb.save(out_path)
