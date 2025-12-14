from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from x4shipqueue.models import EquipmentRow, HullRow


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def autofit_columns(
    ws: Worksheet,
    min_width: int = 10,
    max_width: int = 55,
) -> None:
    """
    Adjust column widths based on content length.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(
            min_width,
            min(max_width, max_len + 2),
        )


# ---------------------------------------------------------------------
# Equipment sheets
# ---------------------------------------------------------------------

def write_equipment_sheet(
    ws: Worksheet,
    rows: List[EquipmentRow],
    max_components: int,
) -> None:
    headers = [
        "Source",
        "Equipment ID",
        "Equipment Name",
        "Race",
        "Size",
        "Mk",
    ]

    for i in range(1, max_components + 1):
        headers += [f"Component {i}", f"Amount {i}"]

    headers.append("Component Count")

    ws.append(headers)

    for r in rows:
        base = [
            r.source,
            r.equipment_id,
            r.equipment_name,
            r.race,
            r.size,
            r.mk,
        ]

        comps: List[object] = []
        for name, amt in r.components[:max_components]:
            comps.extend([name, amt])

        while len(comps) < max_components * 2:
            comps.extend(["", ""])

        ws.append(base + comps + [len(r.components)])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


# ---------------------------------------------------------------------
# Hulls sheet
# ---------------------------------------------------------------------

def write_hulls_sheet(
    ws: Worksheet,
    rows: List[HullRow],
) -> None:
    headers = [
        "Source",
        "Hull ID",
        "Macro ID",
        "Hull Name",
        "Faction",
        "Race",
        "Size",
        "Role",
        "Variant",
        "Crew",
        "Hull HP",
        "Engine Slots",
        "Shield Slots",
        "Weapon Slots",
        "Turret M",
        "Turret L",
        "Price Min",
        "Price Avg",
        "Price Max",
        "Build Time",
        "Energy Cells",
        "Hull Parts",
    ]

    ws.append(headers)

    for h in rows:
        ws.append([
            h.source,
            h.hull_id,
            h.macro_id,
            h.hull_name,
            h.faction,
            h.race,
            h.size,
            h.role,
            h.variant,
            h.crew,
            h.hull_hp,
            h.engine_slots,
            h.shield_slots,
            h.weapon_slots,
            h.turret_m,
            h.turret_l,
            h.price_min,
            h.price_avg,
            h.price_max,
            h.build_time,
            h.energycells,
            h.hullparts,
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws, max_width=70)


# ---------------------------------------------------------------------
# Workbook export
# ---------------------------------------------------------------------

def export_to_excel(
    out_path: Path,
    equipment_by_category: Dict[str, List[EquipmentRow]],
    hull_rows: List[HullRow],
    include_all_equipment: bool,
    max_components: int,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    order = ["Engines", "Thrusters", "Shields", "Weapons", "Turrets"]

    for cat in order:
        ws = wb.create_sheet(cat)
        write_equipment_sheet(
            ws,
            equipment_by_category.get(cat, []),
            max_components=max_components,
        )

    if include_all_equipment:
        all_rows: List[EquipmentRow] = []
        for cat in order:
            all_rows.extend(equipment_by_category.get(cat, []))

        ws = wb.create_sheet("All_Equipment")
        write_equipment_sheet(
            ws,
            all_rows,
            max_components=max_components,
        )

    ws = wb.create_sheet("Hulls")
    write_hulls_sheet(ws, hull_rows)

    wb.save(out_path)
